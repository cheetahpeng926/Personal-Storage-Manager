# -*- coding: utf-8 -*-
"""Personal Storage Manager"""

import json
import logging
import os
import sys
import platform
import shutil
import sqlite3
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.font import Font
from datetime import date, datetime
# Decimal not needed for SQLite

from db import get_connection, init_database

# ── 统一日志 ──
if getattr(sys, 'frozen', False):
    _LOG_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    _LOG_DIR = os.path.dirname(os.path.abspath(__file__))

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(_LOG_DIR, "psm.log"), encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("psm")

# PIL (Pillow is a required dependency)
from PIL import Image, ImageDraw, ImageFont, ImageTk

FONT = "MiSans"
_HEITI = "黑体"

def _detect_font():
    """检测 MiSans 是否可用，不可用时回退到黑体"""
    global FONT
    try:
        import tkinter.font as tkfont
        families = set(tkfont.families())
        if "MiSans" not in families:
            FONT = _HEITI
    except (ImportError, tk.TclError):
        FONT = _HEITI

# ──────────────────────────────────────────────
#  图片目录结构
# ──────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PICTURES_DIR = os.path.join(BASE_DIR, "Pictures")
_POSITION_FILE = os.path.join(BASE_DIR, ".window_pos.json")

def _load_window_pos():
    """读取上次保存的窗口位置，返回 (x, y) 或 None"""
    try:
        with open(_POSITION_FILE, "r") as f:
            pos = json.load(f)
            return int(pos["x"]), int(pos["y"])
    except (IOError, json.JSONDecodeError, KeyError, ValueError) as e:
        logger.debug("读取窗口位置失败: %s", e)
        return None

def _save_window_pos(x, y):
    """保存窗口位置"""
    try:
        with open(_POSITION_FILE, "w") as f:
            json.dump({"x": x, "y": y}, f)
    except IOError as e:
        logger.warning("保存窗口位置失败: %s", e)

IMAGE_FOLDERS = {
    "short_sleeve": {
        "_default": "short-sleeve pic",
    },
    "long_sleeve": {
        "_default": "long-sleeve pic",
    },
    "shorts": {
        "_default": "shorts pic",
    },
    "trousers": {
        "_default": "trousers pic",
    },
    "socks": {
        "_default": "socks pic",
    },
    "shoes": {
        "_default": "shoes pic",
    },
    "car_models": {
        "_default": "models pic",
    },
    "driving_experience": {
        "_default": "driving pic",
    },
    "electronics": {
        "_default": "electronics pic",
    },
}

def ensure_picture_dirs():
    """创建 Pictures 目录结构"""
    for table, cats in IMAGE_FOLDERS.items():
        for folder_name in cats.values():
            path = os.path.join(PICTURES_DIR, folder_name)
            os.makedirs(path, exist_ok=True)


def get_image_folder(table, category=None):
    """根据表和类别返回图片子目录的绝对路径"""
    cats = IMAGE_FOLDERS.get(table, {})
    if category and category in cats:
        folder_name = cats[category]
    elif "_default" in cats:
        folder_name = cats["_default"]
    else:
        folder_name = list(cats.values())[0] if cats else "misc pic"
        os.makedirs(os.path.join(PICTURES_DIR, folder_name), exist_ok=True)
    return os.path.join(PICTURES_DIR, folder_name)


def make_image_filename(brand, model):
    """生成图片文件名：品牌_型号.原扩展名"""
    brand = (brand or "").strip()
    model = (model or "").strip()
    parts = [p for p in [brand, model] if p]
    return "_".join(parts) if parts else "untitled"


# ──────────────────────────────────────────────
#  数据表结构定义
# ──────────────────────────────────────────────
SCHEMA = {
    "short_sleeve": {
        "cn": "短袖",
        "icon": "👕",
        "fields": [
            ("brand",        "品牌",     "text",    True),
            ("model",        "型号",     "text",    True),
            ("color",        "配色",     "text",    True),
            ("article_no",   "货号",     "text",    True),
            ("length_cm",    "衣长",     "decimal", True),
            ("chest_cm",     "胸围",     "decimal", True),
            ("shoulder_cm",  "肩宽",     "decimal", True),
            ("size_label",   "尺码",     "enum",    True,  ["XXS", "XS", "S", "M", "L", "XL", "2XL"]),
            ("purchase_date","购买日期", "date",    True),
            ("image_path",   "图片",     "image",   True),
        ],
    },
    "long_sleeve": {
        "cn": "长袖",
        "icon": "🧥",
        "fields": [
            ("brand",        "品牌",     "text",    True),
            ("model",        "型号",     "text",    True),
            ("color",        "配色",     "text",    True),
            ("article_no",   "货号",     "text",    True),
            ("length_cm",    "衣长",     "decimal", True),
            ("chest_cm",     "胸围",     "decimal", True),
            ("shoulder_cm",  "肩宽",     "decimal", True),
            ("size_label",   "尺码",     "enum",    True,  ["XXS", "XS", "S", "M", "L", "XL", "2XL"]),
            ("purchase_date","购买日期", "date",    True),
            ("image_path",   "图片",     "image",   True),
        ],
    },
    "shorts": {
        "cn": "短裤",
        "icon": "🩳",
        "fields": [
            ("brand",        "品牌",     "text",    True),
            ("model",        "型号",     "text",    True),
            ("color",        "配色",     "text",    True),
            ("article_no",   "货号",     "text",    True),
            ("length_cm",    "裤长",     "decimal", True),
            ("waist_cm",     "腰围",     "decimal", True),
            ("hip_cm",       "臀围",     "decimal", True),
            ("size_label",   "尺码",     "enum",    True,  ["XXS", "XS", "S", "M", "L", "XL", "2XL"]),
            ("purchase_date","购买日期", "date",    True),
            ("image_path",   "图片",     "image",   True),
        ],
    },
    "trousers": {
        "cn": "长裤",
        "icon": "👖",
        "fields": [
            ("brand",        "品牌",     "text",    True),
            ("model",        "型号",     "text",    True),
            ("color",        "配色",     "text",    True),
            ("article_no",   "货号",     "text",    True),
            ("length_cm",    "裤长",     "decimal", True),
            ("waist_cm",     "腰围",     "decimal", True),
            ("hip_cm",       "臀围",     "decimal", True),
            ("size_label",   "尺码",     "enum",    True,  ["XXS", "XS", "S", "M", "L", "XL", "2XL"]),
            ("purchase_date","购买日期", "date",    True),
            ("image_path",   "图片",     "image",   True),
        ],
    },
    "socks": {
        "cn": "袜子",
        "icon": "🧦",
        "fields": [
            ("brand",        "品牌",     "text",    True),
            ("model",        "型号",     "text",    True),
            ("color",        "配色",     "text",    True),
            ("article_no",   "货号",     "text",    True),
            ("size_label",   "尺码",     "text",    True),
            ("purchase_date","购买日期", "date",    True),
            ("image_path",   "图片",     "image",   True),
        ],
    },
    "shoes": {
        "cn": "鞋子",
        "icon": "👟",
        "fields": [
            ("brand",        "品牌",     "text",    True),
            ("model",        "型号",     "text",    True),
            ("color",        "配色",     "text",    True),
            ("article_no",   "货号",     "text",    True),
            ("mm_size",      "mm",      "text",    True),
            ("eur_size",     "EUR",     "text",    True),
            ("purchase_date","购买日期", "date",    True),
            ("image_path",   "图片",     "image",   True),
        ],
    },
    "car_models": {
        "cn": "汽车模型",
        "icon": "🚗",
        "fields": [
            ("scale",        "比例",     "text",    True),
            ("real_brand",   "实车品牌", "text",    True),
            ("real_model",   "实车型号", "text",    True),
            ("color",        "配色",     "text",    True),
            ("model_brand",  "模型品牌", "text",    True),
            ("purchase_date","购买日期", "date",    True),
            ("image_path",   "图片",     "image",   True),
        ],
    },
    "driving_experience": {
        "cn": "驾驶经历",
        "icon": "🏎",
        "fields": [
            ("year",         "年份",     "text",    True),
            ("model",        "型号",     "text",    True),
            ("version",      "版本",     "text",    True),
            ("drivetrain",   "驱动形式", "text",    True),
            ("image_path",   "图片",     "image",   True),
        ],
    },
    "electronics": {
        "cn": "电子设备",
        "icon": "📱",
        "fields": [
            ("category",     "类别",     "enum",    True,  ["手机", "平板", "手表", "耳机"]),
            ("brand",        "品牌",     "text",    True),
            ("model",        "型号",     "text",    True),
            ("color",        "配色",     "text",    True),
            ("release_date", "发布日期", "date",    True),
            ("purchase_date","购买日期", "date",    True),
            ("status",       "状态",     "enum",    True,  ["使用中", "闲置中", "已出售", "已损坏"]),
            ("image_path",   "图片",     "image",   True),
        ],
    },
}

TABLE_ORDER = ["short_sleeve", "long_sleeve", "shorts", "trousers", "socks", "shoes", "car_models", "driving_experience", "electronics"]

# 可排序列（按表名 → 字段名列表）
SORTABLE_COLS = {
    "short_sleeve": ["purchase_date"],
    "long_sleeve":  ["purchase_date"],
    "shorts":       ["purchase_date"],
    "trousers":     ["purchase_date"],
    "socks":        ["purchase_date"],
    "shoes":        ["purchase_date"],
    "car_models":   ["scale", "purchase_date"],
    "driving_experience": ["year"],
    "electronics":  ["category", "purchase_date"],
}

# ──────────────────────────────────────────────
#  Apple 风格配色
# ──────────────────────────────────────────────
COLORS = {
    "bg":            "#F5F5F7",
    "sidebar":       "#FFFFFF",
    "card":          "#FFFFFF",
    "accent":        "#007AFF",
    "accent_hover":  "#0051D5",
    "danger":        "#FF3B30",
    "danger_hover":  "#D63028",
    "success":       "#34C759",
    "text":          "#1D1D1F",
    "text_secondary":"#86868B",
    "separator":     "#E5E5EA",
    "row_hover":     "#F2F2F7",
    "row_selected":  "#E8F0FE",
    "input_bg":      "#FFFFFF",
    "input_border":  "#D2D2D7",
}


# ──────────────────────────────────────────────
#  圆角矩形绘制工具（标准四弧+直线）
# ──────────────────────────────────────────────
def _rounded_rect(canvas, x1, y1, x2, y2, r, **kwargs):
    """在 Canvas 上画标准圆角矩形（圆角用完整椭圆填充，无怪线）"""
    r = min(r, (x2 - x1) / 2, (y2 - y1) / 2)
    fill = kwargs.pop("fill", "")
    outline = kwargs.pop("outline", "")
    tags = kwargs.pop("tags", "")
    items = []
    # 填充：三块矩形 + 四个完整椭圆（覆盖角落，无 pie-slice 怪线）
    items.append(canvas.create_rectangle(x1 + r, y1, x2 - r, y2,
                                         fill=fill, outline="", tags=tags))
    items.append(canvas.create_rectangle(x1, y1 + r, x2, y2 - r,
                                         fill=fill, outline="", tags=tags))
    items.append(canvas.create_oval(x1, y1, x1 + 2*r, y1 + 2*r,
                                    fill=fill, outline="", tags=tags))
    items.append(canvas.create_oval(x2 - 2*r, y1, x2, y1 + 2*r,
                                    fill=fill, outline="", tags=tags))
    items.append(canvas.create_oval(x1, y2 - 2*r, x1 + 2*r, y2,
                                    fill=fill, outline="", tags=tags))
    items.append(canvas.create_oval(x2 - 2*r, y2 - 2*r, x2, y2,
                                    fill=fill, outline="", tags=tags))
    # 边框：四段纯弧线 + 四条直线（如果需要 outline）
    if outline:
        items.append(canvas.create_arc(x1, y1, x1 + 2*r, y1 + 2*r,
                                       start=90, extent=90,
                                       style="arc", outline=outline,
                                       tags=tags))
        items.append(canvas.create_arc(x2 - 2*r, y1, x2, y1 + 2*r,
                                       start=0, extent=90,
                                       style="arc", outline=outline,
                                       tags=tags))
        items.append(canvas.create_arc(x2 - 2*r, y2 - 2*r, x2, y2,
                                       start=270, extent=90,
                                       style="arc", outline=outline,
                                       tags=tags))
        items.append(canvas.create_arc(x1, y2 - 2*r, x1 + 2*r, y2,
                                       start=180, extent=90,
                                       style="arc", outline=outline,
                                       tags=tags))
        items.append(canvas.create_line(x1 + r, y1, x2 - r, y1,
                                        fill=outline, tags=tags))
        items.append(canvas.create_line(x2, y1 + r, x2, y2 - r,
                                        fill=outline, tags=tags))
        items.append(canvas.create_line(x1 + r, y2, x2 - r, y2,
                                        fill=outline, tags=tags))
        items.append(canvas.create_line(x1, y1 + r, x1, y2 - r,
                                        fill=outline, tags=tags))
    return items


# ──────────────────────────────────────────────
#  椭圆滑块自定义滚动条
# ──────────────────────────────────────────────
class OvalScrollbar(tk.Canvas):
    """Canvas 自定义滚动条：椭圆滑块，细轨道"""
    def __init__(self, parent, orient="vertical", command=None, **kw):
        self._orient = orient
        self._command = command
        self._thumb_ratio = 0.0   # 滑块占轨道比例 (0~1)
        self._thumb_pos = 0.0     # 滑块位置比例 (0~1)
        self._drag_start = None
        self._active = False

        parent_bg = parent.cget("bg") if hasattr(parent, "cget") else COLORS["card"]
        if orient == "vertical":
            super().__init__(parent, width=10, bg=parent_bg,
                             highlightthickness=0, cursor="arrow")
        else:
            super().__init__(parent, height=10, bg=parent_bg,
                             highlightthickness=0, cursor="arrow")

        self.bind("<Configure>", self._redraw)
        self.bind("<ButtonPress-1>", self._on_press)
        self.bind("<B1-Motion>", self._on_drag)
        self.bind("<ButtonRelease-1>", self._on_release)
        self.bind("<Enter>", lambda e: self._set_active(True))
        self.bind("<Leave>", lambda e: self._set_active(False))

    def set(self, first, last):
        """由关联控件的 yscrollcommand/xscrollcommand 调用"""
        self._thumb_pos = float(first)
        self._thumb_ratio = float(last) - float(first)
        self._redraw()

    def _set_active(self, active):
        self._active = active
        self._redraw()

    def _redraw(self, event=None):
        self.delete("thumb")
        if self._orient == "vertical":
            w = self.winfo_width()
            h = self.winfo_height()
        else:
            w = self.winfo_width()
            h = self.winfo_height()

        if w < 4 or h < 4:
            return

        # 轨道
        if self._orient == "vertical":
            track_x = w // 2
            self.create_line(track_x, 4, track_x, h - 4,
                             fill="#E8E8EC", width=1, tags="thumb")
        else:
            track_y = h // 2
            self.create_line(4, track_y, w - 4, track_y,
                             fill="#E8E8EC", width=1, tags="thumb")

        if self._thumb_ratio >= 1.0:
            return

        min_thumb = 32
        color = "#C0C0C4" if not self._active else "#909094"

        if self._orient == "vertical":
            thumb_h = max(min_thumb, self._thumb_ratio * h)
            thumb_top = self._thumb_pos * (h - thumb_h)
            cx = w / 2
            ry = thumb_h / 2
            rx = w / 2 - 1
            cy = thumb_top + ry
            self.create_oval(cx - rx, cy - ry, cx + rx, cy + ry,
                             fill=color, outline="", tags="thumb")
        else:
            thumb_w = max(min_thumb, self._thumb_ratio * w)
            thumb_left = self._thumb_pos * (w - thumb_w)
            cy = h / 2
            rx = thumb_w / 2
            ry = h / 2 - 1
            cx = thumb_left + rx
            self.create_oval(cx - rx, cy - ry, cx + rx, cy + ry,
                             fill=color, outline="", tags="thumb")

    def _on_press(self, event):
        if self._orient == "vertical":
            h = self.winfo_height()
            thumb_h = max(28, self._thumb_ratio * h)
            thumb_top = self._thumb_pos * (h - thumb_h)
            if thumb_top <= event.y <= thumb_top + thumb_h:
                self._drag_start = (event.y, self._thumb_pos)
            else:
                # 点击轨道跳转
                ratio = event.y / h
                if self._command:
                    self._command("moveto", str(ratio))
        else:
            w = self.winfo_width()
            thumb_w = max(28, self._thumb_ratio * w)
            thumb_left = self._thumb_pos * (w - thumb_w)
            if thumb_left <= event.x <= thumb_left + thumb_w:
                self._drag_start = (event.x, self._thumb_pos)
            else:
                ratio = event.x / w
                if self._command:
                    self._command("moveto", str(ratio))

    def _on_drag(self, event):
        if self._drag_start is None:
            return
        if self._orient == "vertical":
            h = self.winfo_height()
            thumb_h = max(28, self._thumb_ratio * h)
            dy = event.y - self._drag_start[0]
            new_pos = self._drag_start[1] + dy / max(1, h - thumb_h)
            new_pos = max(0.0, min(1.0, new_pos))
            if self._command:
                self._command("moveto", str(new_pos))
        else:
            w = self.winfo_width()
            thumb_w = max(28, self._thumb_ratio * w)
            dx = event.x - self._drag_start[0]
            new_pos = self._drag_start[1] + dx / max(1, w - thumb_w)
            new_pos = max(0.0, min(1.0, new_pos))
            if self._command:
                self._command("moveto", str(new_pos))

    def _on_release(self, event):
        self._drag_start = None


# ──────────────────────────────────────────────
#  圆角按钮
# ──────────────────────────────────────────────
class RoundedButton(tk.Canvas):
    def __init__(self, parent, text="", command=None, width=90, height=34,
                 bg=COLORS["accent"], fg="#FFFFFF", hover_bg=None, radius=7,
                 font_size=11, icon="", **kw):
        # 继承父容器背景色，避免白底露出
        parent_bg = kw.pop("canvas_bg", None) or parent.cget("bg") if hasattr(parent, "cget") else COLORS["bg"]
        super().__init__(parent, width=width, height=height,
                         bg=parent_bg, highlightthickness=0, **kw)
        self._bg = bg
        self._hover_bg = hover_bg or COLORS["accent_hover"]
        self._fg = fg
        self._radius = radius
        self._command = command
        self._text = text
        self._icon = icon
        self._font_size = font_size
        self._width = width
        self._height = height

        self.bind("<Enter>", lambda e: self._draw(self._hover_bg))
        self.bind("<Leave>", lambda e: self._draw(self._bg))
        self.bind("<Button-1>", lambda e: self._command() if self._command else None)

        self._draw(bg)

    def _draw(self, fill_color):
        self.delete("all")
        _rounded_rect(self, 0, 0, self._width, self._height,
                      self._radius, fill=fill_color, outline="")
        label = f"{self._icon} {self._text}" if self._icon else self._text
        self.create_text(self._width//2, self._height//2, text=label,
                         fill=self._fg, font=(FONT, self._font_size, ""))


# ──────────────────────────────────────────────
#  侧边栏按钮
# ──────────────────────────────────────────────
class SidebarButton(tk.Canvas):
    def __init__(self, parent, icon, text, command=None,
                 selected=False, width=200, height=42, **kw):
        super().__init__(parent, width=width, height=height,
                         bg=COLORS["sidebar"], highlightthickness=0, **kw)
        self._icon = icon
        self._text = text
        self._command = command
        self._selected = selected
        self._width = width
        self._height = height

        self.bind("<Enter>", lambda e: self._draw(hover=True))
        self.bind("<Leave>", lambda e: self._draw())
        self.bind("<Button-1>", lambda e: self._command() if self._command else None)

        self._draw()

    def _draw(self, hover=False):
        self.delete("all")
        w, h = self._width, self._height
        if self._selected:
            _rounded_rect(self, 6, 3, w-6, h-3, 7,
                          fill="#CCE4FF", outline="#A8CFFF")
            text_fill = "#0051D5"
        elif hover:
            _rounded_rect(self, 6, 3, w-6, h-3, 7,
                          fill="#EAF2FF", outline="#D6E8FF")
            text_fill = COLORS["text"]
        else:
            text_fill = COLORS["text"]

        self.create_text(35, h//2, text=self._icon, font=(FONT, 16), anchor="center")
        self.create_text(60, h//2, text=self._text, fill=text_fill,
                         font=(FONT, 13, ""), anchor="w")

    def set_selected(self, selected):
        self._selected = selected
        self._draw()


# ──────────────────────────────────────────────
#  主应用
# ──────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Personal Storage Manager")
        self.geometry("1235x763")
        self.minsize(1060, 600)
        self.configure(bg=COLORS["bg"])

        # 窗口位置：记住上次位置
        pos = _load_window_pos()
        if pos:
            self.geometry(f"+{pos[0]}+{pos[1]}")
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # 窗口图标：用 PIL 渲染 📦 emoji，失败则保持默认
        try:
            _icon_img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
            _draw = ImageDraw.Draw(_icon_img)
            _emoji_font = None
            for _fp in [
                "/usr/share/fonts/truetype/noto/NotoColorEmoji.ttf",
                "/usr/share/fonts/google-noto-emoji/NotoColorEmoji.ttf",
                "/usr/share/fonts/noto-emoji/NotoColorEmoji.ttf",
                "/System/Library/Fonts/Apple Color Emoji.ttc",
                "C:\\Windows\\Fonts\\seguiemj.ttf",
            ]:
                if os.path.isfile(_fp):
                    _emoji_font = ImageFont.truetype(_fp, 52)
                    break
            if _emoji_font:
                _bbox = _emoji_font.getbbox("📦")
                _tw = _bbox[2] - _bbox[0]
                _th = _bbox[3] - _bbox[1]
                _x = (64 - _tw) // 2 - _bbox[0]
                _y = (64 - _th) // 2 - _bbox[1]
                _draw.text((_x, _y), "📦", font=_emoji_font, embedded_color=True)
            else:
                # fallback：画一个简易包裹图标
                _draw.rounded_rectangle([10, 16, 54, 52], radius=4, fill="#C8956A", outline="#8B6543", width=2)
                _draw.rectangle([22, 10, 42, 22], fill="#D4A574", outline="#8B6543", width=2)
                _draw.line([32, 10, 32, 52], fill="#8B6543", width=2)
                _draw.line([10, 34, 54, 34], fill="#8B6543", width=2)
            self._icon_photo = ImageTk.PhotoImage(_icon_img)
            self.iconphoto(True, self._icon_photo)
            if platform.system() == "Windows":
                import ctypes
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("psm.psm.psm")
        except (OSError, tk.TclError) as e:
            logger.debug("设置应用图标失败: %s", e)

        # 字体检测：MiSans 不可用时回退到黑体
        _detect_font()

        # 下拉框弹出列表背景白色
        self.option_add('*TCombobox*Listbox.background', COLORS["input_bg"])
        self.option_add('*TCombobox*Listbox.foreground', COLORS["text"])
        self.option_add('*TCombobox*Listbox.selectBackground', COLORS["accent"])
        self.option_add('*TCombobox*Listbox.selectForeground', '#FFFFFF')

        ensure_picture_dirs()

        try:
            init_database()
        except sqlite3.Error as e:
            messagebox.showerror("数据库错误",
                                 f"数据库初始化失败\n\n{e}")
            self.destroy()
            return

        self._conn = get_connection()

        self._current_table = "short_sleeve"
        self._sidebar_buttons = {}
        self._photo_cache = []   # 防止 PhotoImage 被 GC
        self._current_img_path = None  # 当前预览图片路径，点击可打开
        self._sort_col = None    # 当前排序列
        self._sort_dir = 0       # 0=默认 1=升序 2=降序
        self._build_ui()
        self._refresh_table()

    def _on_close(self):
        """关闭窗口时保存位置并断开数据库"""
        self.update_idletasks()
        _save_window_pos(self.winfo_x(), self.winfo_y())
        try:
            self._conn.close()
        except Exception as e:
            logger.warning("关闭数据库连接失败: %s", e)
        self.destroy()

    # ─── UI 构建 ───────────────────────────────
    def _build_ui(self):
        main = tk.Frame(self, bg=COLORS["bg"])
        main.pack(fill="both", expand=True)

        # ═══ 左侧边栏 + 圆角接缝 + 内容区 ═══
        main.grid_columnconfigure(2, weight=1)
        main.grid_rowconfigure(0, weight=1)

        sidebar = tk.Frame(main, bg=COLORS["sidebar"], width=220)
        sidebar.grid(row=0, column=0, sticky="ns")
        sidebar.pack_propagate(False)

        # 简洁分割线（单条浅蓝）
        seam = tk.Canvas(main, width=1, bg=COLORS["bg"],
                         highlightthickness=0)
        seam.grid(row=0, column=1, sticky="ns")

        def _draw_seam(event=None):
            seam.delete("all")
            sh = seam.winfo_height()
            if sh < 10:
                return
            seam.create_rectangle(0, 0, 1, sh, fill="#D0E0F0", outline="")

        seam.bind("<Configure>", _draw_seam)

        content = tk.Frame(main, bg=COLORS["bg"])
        content.grid(row=0, column=2, sticky="nsew")

        tk.Frame(sidebar, bg=COLORS["separator"], height=1).pack(fill="x", side="top")

        # Logo
        logo_frame = tk.Frame(sidebar, bg=COLORS["sidebar"])
        logo_frame.pack(fill="x", pady=10)
        tk.Label(logo_frame, text="📦", font=(FONT, 28),
                 bg=COLORS["sidebar"]).pack()
        tk.Label(logo_frame, text="Personal Storage",
                 font=(FONT, 14, "bold"), bg=COLORS["sidebar"],
                 fg=COLORS["text"]).pack(pady=(6, 0))
        tk.Label(logo_frame, text="Manager",
                 font=(FONT, 14), bg=COLORS["sidebar"],
                 fg=COLORS["text"]).pack()
        tk.Label(logo_frame, text="Made by PWR",
                 font=(FONT, 9), bg=COLORS["sidebar"],
                 fg=COLORS["text_secondary"]).pack(pady=(8, 0))
        tk.Label(logo_frame, text="Built with Xiaomi MiMo V2",
                 font=(FONT, 9), bg=COLORS["sidebar"],
                 fg=COLORS["text_secondary"]).pack()

        tk.Frame(sidebar, bg=COLORS["separator"], height=1).pack(fill="x", padx=16, pady=0)

        tk.Label(sidebar, text="分  类", font=(FONT, 10),
                 bg=COLORS["sidebar"], fg=COLORS["text_secondary"],
                 anchor="w").pack(fill="x", padx=20, pady=(10, 4))

        for table in TABLE_ORDER:
            schema = SCHEMA[table]
            btn = SidebarButton(
                sidebar,
                icon=schema["icon"],
                text=schema["cn"],
                command=lambda t=table: self._switch_table(t),
                selected=(table == "short_sleeve"),
            )
            btn.pack(fill="x", padx=10, pady=2)
            self._sidebar_buttons[table] = btn

        # 侧边栏底部：导出按钮
        footer = tk.Frame(sidebar, bg=COLORS["sidebar"])
        footer.pack(side="bottom", fill="x", padx=12, pady=(0, 16))
        RoundedButton(footer, text="导出", icon="📤",
                      command=self._export_xlsx,
                      bg="#E8F8EE", fg="#1B8A4A",
                      hover_bg="#D0F0DC", width=200, height=36, radius=7).pack()

        # ── 顶栏 ──
        header = tk.Frame(content, bg=COLORS["bg"])
        header.pack(fill="x", padx=24, pady=(20, 0))

        title_frame = tk.Frame(header, bg=COLORS["bg"])
        title_frame.pack(side="left")

        self._title_icon = tk.Label(
            title_frame, text=SCHEMA['short_sleeve']['icon'],
            font=(FONT, 22), bg=COLORS["bg"], fg=COLORS["text"])
        self._title_icon.pack(side="left")
        self._title_text = tk.Label(
            title_frame, text=SCHEMA['short_sleeve']['cn'],
            font=(FONT, 22, "bold"), bg=COLORS["bg"], fg=COLORS["text"])
        self._title_text.pack(side="left", padx=(10, 0))

        self._count_label = tk.Label(
            title_frame, text="共 0 条记录",
            font=(FONT, 11), bg=COLORS["bg"], fg=COLORS["text_secondary"])
        self._count_label.pack(side="bottom", anchor="sw")

        right_frame = tk.Frame(header, bg=COLORS["bg"])
        right_frame.pack(side="right")

        # 搜索框（圆角）
        search_canvas = tk.Canvas(right_frame, width=210, height=34,
                                  bg=COLORS["bg"], highlightthickness=0)
        search_canvas.pack(side="left", padx=(0, 4), pady=(2, 0))

        _rounded_rect(search_canvas, 1, 1, 209, 33, 7,
                      fill=COLORS["input_bg"], outline=COLORS["input_border"],
                      tags="search_bg")
        search_canvas.create_text(14, 17, text="搜索", font=(FONT, 10),
                                  fill=COLORS["text_secondary"],
                                  anchor="w", tags="search_icon")
        self.var_search = tk.StringVar()
        search_entry = tk.Entry(search_canvas, textvariable=self.var_search,
                                font=(FONT, 12), bg=COLORS["input_bg"],
                                fg=COLORS["text"], relief="flat", width=14,
                                insertbackground=COLORS["accent"],
                                borderwidth=0)
        search_canvas.create_window(42, 17, window=search_entry, anchor="w")

        def _search_focus_in(e):
            search_canvas.delete("search_bg")
            _rounded_rect(search_canvas, 1, 1, 209, 33, 7,
                          fill=COLORS["input_bg"], outline=COLORS["accent"],
                          tags="search_bg")
            search_canvas.lower("search_bg")
        def _search_focus_out(e):
            search_canvas.delete("search_bg")
            _rounded_rect(search_canvas, 1, 1, 209, 33, 7,
                          fill=COLORS["input_bg"], outline=COLORS["input_border"],
                          tags="search_bg")
            search_canvas.lower("search_bg")
        search_entry.bind("<FocusIn>", _search_focus_in)
        search_entry.bind("<FocusOut>", _search_focus_out)

        # 操作按钮
        btn_frame = tk.Frame(right_frame, bg=COLORS["bg"])
        btn_frame.pack(side="left")

        RoundedButton(btn_frame, text="新增", icon="＋",
                      command=self._add_record,
                      bg="#CCE4FF", fg="#0051D5",
                      hover_bg="#B8D8FF",
                      width=90, height=34, radius=7).pack(side="left", padx=4)
        RoundedButton(btn_frame, text="编辑", icon="✎",
                      command=self._edit_record,
                      bg="#EAF2FF", fg="#0051D5",
                      hover_bg="#D6E8FF", width=90, height=34, radius=7).pack(side="left", padx=4)
        RoundedButton(btn_frame, text="删除", icon="🗑",
                      command=self._delete_record,
                      bg="#EAF2FF", fg="#D63028",
                      hover_bg="#FFE5E3", width=90, height=34, radius=7).pack(side="left", padx=4)

        # ── 主体：左表格 + 右图片预览 ──
        body = tk.Frame(content, bg=COLORS["bg"])
        body.pack(fill="both", expand=True, padx=24, pady=(12, 24))
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(0, weight=1)

        # 表格卡片（Frame + 细边框，干净无双层框）
        self._table_container = tk.Frame(body, bg=COLORS["card"],
                                         highlightbackground=COLORS["separator"],
                                         highlightthickness=1)
        self._table_container.grid(row=0, column=0, sticky="nsew")
        self._table_container.grid_columnconfigure(0, weight=1)
        self._table_container.grid_rowconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Apple.Treeview",
                        background=COLORS["card"],
                        foreground=COLORS["text"],
                        fieldbackground=COLORS["card"],
                        borderwidth=0, relief="flat",
                        font=(FONT, 12), rowheight=52,
                        indent=0)
        style.configure("Apple.Treeview.Heading",
                        background=COLORS["card"],
                        foreground=COLORS["text_secondary"],
                        borderwidth=0, relief="flat",
                        padding=(8, 6),
                        font=(FONT, 11, "bold"))
        style.map("Apple.Treeview",
                  background=[("selected", COLORS["row_selected"])],
                  foreground=[("selected", COLORS["text"])])
        style.map("Apple.Treeview.Heading",
                  background=[("active", COLORS["separator"])])

        # 表单下拉框样式：白色背景，和输入框一致
        style.configure("Form.TCombobox",
                        fieldbackground=COLORS["input_bg"],
                        background=COLORS["input_bg"],
                        bordercolor=COLORS["input_border"],
                        arrowcolor=COLORS["text"],
                        font=(FONT, 12))
        style.map("Form.TCombobox",
                  fieldbackground=[("readonly", COLORS["input_bg"])],
                  background=[("readonly", COLORS["input_bg"])])

        self.tree = ttk.Treeview(self._table_container, show="headings",
                                 selectmode="browse", style="Apple.Treeview")
        self.tree.column("#0", width=0, minwidth=0, stretch=False)
        # 垂直滚动条：隐藏但功能保留（供鼠标滚轮驱动）
        vsb = OvalScrollbar(self._table_container, orient="vertical",
                            command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        # 不要水平滚动条，列宽自动适配

        # 鼠标滚轮支持
        def _on_mousewheel(event):
            _hide_tooltip()
            self.tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.tree.bind("<MouseWheel>", _on_mousewheel)
        # Linux 下的滚轮事件
        self.tree.bind("<Button-4>", lambda e: (_hide_tooltip(), self.tree.yview_scroll(-3, "units")))
        self.tree.bind("<Button-5>", lambda e: (_hide_tooltip(), self.tree.yview_scroll(3, "units")))

        self.tree.tag_configure("odd", background=COLORS["card"])
        self.tree.tag_configure("even", background="#FAFAFA")
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<ButtonRelease-1>", self._on_select)
        self.tree.bind("<Double-1>", lambda e: self._edit_record())

        # 搜索框实时过滤（tree 创建后再绑定，避免初始化时触发刷新）
        self.var_search.trace_add("write", lambda *_: self._refresh_table())

        # 窗口大小变化时自动调整列宽
        def _on_tree_resize(event=None):
            cols = self.tree["columns"]
            if not cols:
                return
            w = self._table_container.winfo_width()
            if w < 50:
                return
            col_w = max(60, w // len(cols))
            for cid in cols:
                self.tree.column(cid, width=col_w)
        self._table_container.bind("<Configure>", _on_tree_resize)

        # ── 单元格 Tooltip（悬停显示完整内容）──
        self._tooltip_win = None
        def _hide_tooltip():
            if self._tooltip_win:
                self._tooltip_win.destroy()
                self._tooltip_win = None
        self._hide_tooltip = _hide_tooltip

        def _show_tooltip(event):
            row_id = self.tree.identify_row(event.y)
            col_id = self.tree.identify_column(event.x)
            if not row_id or not col_id:
                _hide_tooltip()
                return
            col_idx = int(col_id.replace("#", "")) - 1
            cols = self.tree["columns"]
            if col_idx < 0 or col_idx >= len(cols):
                _hide_tooltip()
                return
            values = self.tree.item(row_id, "values")
            val_idx = col_idx  # 直接使用列索引，无需偏移
            if not values or val_idx < 0 or val_idx >= len(values):
                _hide_tooltip()
                return
            text = str(values[val_idx])
            # 检查文字是否被截断（列宽不够时才显示 tooltip）
            col_width = self.tree.column(cols[col_idx], "width")
            f = Font(family=FONT, size=12)
            text_width = f.measure(text)
            if text_width <= col_width - 8:
                _hide_tooltip()
                return
            _hide_tooltip()
            tip = tk.Toplevel(self)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{event.x_root + 16}+{event.y_root + 10}")
            tip.configure(bg="#FFFFDD")
            lbl = tk.Label(tip, text=text, font=(FONT, 11),
                           bg="#FFFFDD", fg="#333333",
                           justify="left", wraplength=360,
                           padx=8, pady=4)
            lbl.pack()
            self._tooltip_win = tip

        self.tree.bind("<Motion>", _show_tooltip)
        self.tree.bind("<Leave>", lambda e: _hide_tooltip())

        # ── 图片预览面板（标准圆角矩形）──
        preview_panel = tk.Canvas(body, bg=COLORS["bg"],
                                  highlightthickness=0, width=230)
        preview_panel.grid(row=0, column=1, sticky="ns", padx=(12, 0))

        def _draw_preview_bg(event=None):
            preview_panel.delete("bg_shape")
            pw = preview_panel.winfo_width()
            ph = preview_panel.winfo_height()
            if pw < 20 or ph < 20:
                return
            _rounded_rect(preview_panel, 1, 1, pw-1, ph-1, 7,
                          fill=COLORS["card"],
                          outline=COLORS["separator"],
                          tags="bg_shape")
            preview_panel.lower("bg_shape")

        preview_panel.bind("<Configure>", _draw_preview_bg)

        tk.Label(preview_panel, text="图片预览", font=(FONT, 12, "bold"),
                 bg=COLORS["card"], fg=COLORS["text"]).place(relx=0.5, y=28, anchor="center")

        self._preview_label = tk.Label(
            preview_panel, text="选中记录后\n显示图片",
            font=(FONT, 11), bg=COLORS["card"],
            fg=COLORS["text_secondary"], compound="top", cursor="hand2")
        self._preview_label.place(x=5, y=50, width=220, height=340, anchor="nw")
        self._preview_label.bind("<Button-1>", self._open_preview_image)

        # 图片操作按钮（上下排列，底部对齐）
        img_btn_frame = tk.Frame(preview_panel, bg=COLORS["card"])
        img_btn_frame.place(relx=0.5, rely=1.0, y=-10, anchor="s")

        RoundedButton(img_btn_frame, text="更换图片", icon="🖼",
                      command=self._change_image,
                      bg="#CCE4FF", fg="#0051D5",
                      hover_bg="#B8D8FF",
                      width=200, height=32, font_size=10, radius=7).pack(pady=(0, 4))
        RoundedButton(img_btn_frame, text="删除图片", icon="✕",
                      command=self._remove_image,
                      bg="#EAF2FF", fg="#D63028",
                      hover_bg="#FFE5E3",
                      width=200, height=32,
                      font_size=10, radius=7).pack()

    # ─── 图片预览 ──────────────────────────────
    def _on_select(self, event=None):
        """点击/选中记录后加载图片预览"""
        self._hide_tooltip()
        sel = self.tree.selection()
        if not sel:
            return
        tags = self.tree.item(sel[0], "tags")
        rid = int(tags[0]) if tags else None
        if rid is None:
            return
        self._load_preview_for(rid)

    def _load_preview_for(self, rid):
        """查询数据库并显示图片预览"""
        try:
            table = self._get_table()
            cur = self._conn.cursor()
            cur.execute(f"SELECT image_path FROM `{table}` WHERE id = ?", (rid,))
            row = cur.fetchone()
            cur.close()

            img_path = row["image_path"] if row else None
            self._show_preview(img_path)
        except Exception as e:
            self._preview_label.configure(image="", text=f"预览出错\n{e}")

    def _show_preview(self, img_path):
        self._photo_cache.clear()
        self._current_img_path = img_path if img_path and os.path.isfile(img_path) else None
        if self._current_img_path:
            try:
                img = Image.open(img_path)
                img.thumbnail((236, 236))
                photo = ImageTk.PhotoImage(img)
                self._photo_cache.append(photo)
                self._preview_label.configure(image=photo, text="")
                self._preview_label.update_idletasks()
                return
            except Exception:
                pass

            # fallback: 无 PIL 时用 tkinter PhotoImage（仅支持 png/gif）
            try:
                ext = os.path.splitext(img_path)[1].lower()
                if ext in (".png", ".gif"):
                    photo = tk.PhotoImage(file=img_path)
                    self._photo_cache.append(photo)
                    self._preview_label.configure(image=photo, text="")
                    self._preview_label.update_idletasks()
                    return
            except Exception:
                pass

            self._preview_label.configure(image="",
                                          text=f"无法预览\n{os.path.basename(img_path)}")
        else:
            self._preview_label.configure(image="",
                                          text="选中记录后\n显示图片")
        self._preview_label.update_idletasks()

    def _open_preview_image(self, event=None):
        """点击预览图片，用系统默认程序打开"""
        if not self._current_img_path:
            return
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(self._current_img_path)
            elif system == "Darwin":
                subprocess.Popen(["open", self._current_img_path])
            else:
                subprocess.Popen(["xdg-open", self._current_img_path])
        except (OSError, subprocess.SubprocessError) as e:
            logger.warning("打开图片失败: %s", e)

    def _change_image(self):
        """为选中记录添加/更换图片"""
        table = self._get_table()
        rid = self._get_selected_id()
        if rid is None:
            return

        cur = self._conn.cursor()
        cur.execute(f"SELECT * FROM `{table}` WHERE id = ?", (rid,))
        row = cur.fetchone()
        cur.close()

        if not row:
            return

        # 获取类别（用于决定存到哪个子目录）
        r = dict(row)
        category = r.get("category") or None
        brand = r.get("brand") or r.get("real_brand") or r.get("model_brand") or ""
        model = r.get("model") or r.get("real_model") or ""

        # 选文件
        src = filedialog.askopenfilename(
            title="选择图片",
            filetypes=[
                ("图片文件", "*.jpg *.jpeg *.png *.gif *.bmp *.webp"),
                ("所有文件", "*.*"),
            ])
        if not src:
            return

        # 确定目标目录和文件名
        folder = get_image_folder(table, category)
        ext = os.path.splitext(src)[1].lower()
        if not ext:
            ext = ".jpg"
        filename = make_image_filename(brand, model) + ext
        dst = os.path.join(folder, filename)

        # 记录旧图片路径，稍后清理
        old_path = r.get("image_path")

        # 如果同名文件已存在，加序号
        if os.path.abspath(dst) == os.path.abspath(src):
            pass  # 就是原文件
        else:
            base, fext = os.path.splitext(filename)
            counter = 1
            while os.path.exists(dst) and os.path.abspath(dst) != os.path.abspath(old_path or ""):
                filename = f"{base}_{counter}{fext}"
                dst = os.path.join(folder, filename)
                counter += 1
            shutil.copy2(src, dst)

        # 更新数据库
        cur = self._conn.cursor()
        cur.execute(f"UPDATE `{table}` SET image_path = ? WHERE id = ?",
                    (dst, rid))
        self._conn.commit()
        cur.close()

        # 清理旧图片文件（仅当旧图片不是新图片时才删除）
        if old_path and os.path.isfile(old_path) and os.path.abspath(old_path) != os.path.abspath(dst):
            try:
                os.remove(old_path)
            except OSError:
                pass

        self._show_preview(dst)
        self._refresh_table()

    def _remove_image(self):
        """删除选中记录的图片（同时清理物理文件）"""
        table = self._get_table()
        rid = self._get_selected_id()
        if rid is None:
            return

        # 先查出旧图片路径
        cur = self._conn.cursor()
        cur.execute(f"SELECT image_path FROM `{table}` WHERE id = ?", (rid,))
        row = cur.fetchone()
        old_path = row["image_path"] if row else None

        cur.execute(f"UPDATE `{table}` SET image_path = NULL WHERE id = ?", (rid,))
        self._conn.commit()
        cur.close()

        # 删除物理文件
        if old_path and os.path.isfile(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

        self._show_preview(None)
        self._refresh_table()

    # ─── 切换分类 ──────────────────────────────
    def _switch_table(self, table):
        self._current_table = table
        self._sort_col = None
        self._sort_dir = 0
        for t, btn in self._sidebar_buttons.items():
            btn.set_selected(t == table)

        schema = SCHEMA[table]
        self._title_icon.configure(text=schema['icon'])
        self._title_text.configure(text=schema['cn'])
        self._hide_tooltip()
        self.var_search.set("")
        self._show_preview(None)

    # ─── 数据加载 ──────────────────────────────
    def _get_table(self):
        return self._current_table

    def _refresh_table(self):
        table = self._get_table()
        schema = SCHEMA[table]
        sortable = SORTABLE_COLS.get(table, [])

        self.tree.delete(*self.tree.get_children())
        self._photo_cache.clear()

        # 表格中不显示 image_path 列，但保留 id
        display_fields = [f for f in schema["fields"]
                          if f[0] not in ("id", "image_path")]
        col_names = [f[1] for f in display_fields]
        col_ids = [f[0] for f in display_fields]

        # 设置列
        self.tree["columns"] = col_ids

        n_cols = len(col_ids)
        self.update_idletasks()
        container_w = self._table_container.winfo_width()
        if container_w < 50:
            container_w = 900
        col_width = max(70, container_w // n_cols)
        for cid, cname in zip(col_ids, col_names):
            # 可排序列加箭头指示
            if cid in sortable:
                if self._sort_col == cid and self._sort_dir == 1:
                    heading_text = f"{cname} ▲"
                elif self._sort_col == cid and self._sort_dir == 2:
                    heading_text = f"{cname} ▼"
                else:
                    heading_text = f"{cname} ↕"
                self.tree.heading(cid, text=heading_text, anchor="center",
                                  command=lambda c=cid: self._on_heading_click(c))
            else:
                self.tree.heading(cid, text=cname, anchor="center")
            self.tree.column(cid, width=col_width, anchor="center", minwidth=60,
                             stretch=True)

        cur = self._conn.cursor()

        # 构建 ORDER BY
        if self._sort_col and self._sort_dir in (1, 2):
            direction = "ASC" if self._sort_dir == 1 else "DESC"
            order_by = f"`{self._sort_col}` {direction}, `id` DESC"
        else:
            order_by = "`id` DESC"

        search = self.var_search.get().strip()
        text_fields = [f[0] for f in schema["fields"] if f[2] == "text"]
        if search and text_fields:
            conditions = " OR ".join(f"`{f}` LIKE ?" for f in text_fields)
            params = [f"%{search}%"] * len(text_fields)
            cur.execute(f"SELECT * FROM `{table}` WHERE {conditions} ORDER BY {order_by}", params)
        else:
            cur.execute(f"SELECT * FROM `{table}` ORDER BY {order_by}")

        rows = cur.fetchall()
        cur.close()

        for idx, row in enumerate(rows):
            values = []
            for f in display_fields:
                val = row[f[0]]
                if val is None:
                    values.append("—")
                elif isinstance(val, (int, float)):
                    values.append(str(val))
                elif hasattr(val, "strftime"):
                    values.append(val.strftime("%Y.%m.%d"))
                else:
                    values.append(str(val))
            tag = "even" if idx % 2 == 0 else "odd"
            # 直接插入数据，不使用任何占位符
            self.tree.insert("", "end", values=values,
                             tags=(str(row["id"]), tag))

        self._count_label.configure(text=f"共 {len(rows)} 条记录")

    def _on_heading_click(self, col):
        """点击可排序表头：升序 → 降序 → 恢复默认"""
        if self._sort_col == col:
            # 循环：1→2→0
            self._sort_dir = {1: 2, 2: 0}.get(self._sort_dir, 1)
            if self._sort_dir == 0:
                self._sort_col = None
        else:
            self._sort_col = col
            self._sort_dir = 1
        self._refresh_table()

    # ─── 弹窗表单 ──────────────────────────────
    def _form_dialog(self, title, table, initial=None):
        schema = SCHEMA[table]
        fields = [f for f in schema["fields"]
                  if f[0] not in ("id", "created_at", "image_path")]

        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.configure(bg=COLORS["card"])
        dlg.geometry("440x560")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()

        # 先隐藏窗口，计算好位置后再显示，避免左上角闪烁
        dlg.withdraw()
        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 440) // 2
        y = self.winfo_y() + (self.winfo_height() - 560) // 2
        dlg.geometry(f"+{max(x,0)}+{max(y,0)}")
        dlg.deiconify()

        result = {}

        tk.Label(dlg, text=title, font=(FONT, 18, "bold"),
                 bg=COLORS["card"], fg=COLORS["text"]).pack(pady=(24, 16))

        # 表单区：白色背景
        card = tk.Frame(dlg, bg=COLORS["card"])
        card.pack(fill="x", padx=32)

        entries = {}

        for i, (fname, cn_name, ftype, required, *extra) in enumerate(fields):
            row = tk.Frame(card, bg=COLORS["card"])
            row.pack(fill="x", padx=0, pady=(10 if i == 0 else 4, 4))

            lbl = tk.Label(row, text=cn_name,
                           font=(FONT, 12), bg=COLORS["card"],
                           fg=COLORS["text"],
                           width=8, anchor="w")

            lbl.pack(side="left")

            if ftype == "enum":
                options = extra[0] if extra else []
                var = tk.StringVar()
                if initial and initial.get(fname):
                    var.set(str(initial[fname]))
                elif options:
                    var.set(options[0])

                widget = ttk.Combobox(row, textvariable=var,
                                      values=options, state="readonly",
                                      font=(FONT, 12), style="Form.TCombobox",
                                      justify="center")
                widget.pack(side="left", padx=(6, 0), fill="x", expand=True)
                entries[fname] = ("enum", var)
            else:
                var = tk.StringVar()
                if initial and initial.get(fname):
                    val = initial[fname]
                    if isinstance(val, (int, float)):
                        var.set(str(val))
                    elif hasattr(val, "strftime"):
                        var.set(val.strftime("%Y.%m.%d"))
                    else:
                        var.set(str(val))

                entry = tk.Entry(row, textvariable=var,
                                 font=(FONT, 12), bg=COLORS["input_bg"],
                                 fg=COLORS["text"],
                                 relief="flat", justify="center",
                                 highlightbackground=COLORS["input_border"],
                                 highlightcolor=COLORS["accent"],
                                 highlightthickness=1,
                                 insertbackground=COLORS["accent"])
                entry.pack(side="left", padx=(6, 0), fill="x", expand=True)

                placeholders = {"date": "YYYY.MM.DD"}
                placeholder = placeholders.get(ftype, "")

                if not initial or not var.get():
                    entry.configure(fg=COLORS["text_secondary"])
                    entry.insert(0, placeholder)

                    def make_handlers(e_entry=entry, e_var=var, e_ph=placeholder):
                        def on_in(evt):
                            if e_entry.get() == e_ph:
                                e_entry.delete(0, "end")
                                e_entry.configure(fg=COLORS["text"])
                        def on_out(evt):
                            if not e_entry.get():
                                e_entry.insert(0, e_ph)
                                e_entry.configure(fg=COLORS["text_secondary"])
                        return on_in, on_out

                    on_in, on_out = make_handlers()
                    entry.bind("<FocusIn>", on_in)
                    entry.bind("<FocusOut>", on_out)

                # ---------- 自动格式化：日期 8位数字 → YYYY.MM.DD ----------
                if ftype == "date":
                    def _auto_date(evt, e_var=var):
                        raw = e_var.get().strip()
                        if raw == placeholders.get("date", ""):
                            return
                        # 纯8位数字 → 自动加点
                        if len(raw) == 8 and raw.isdigit():
                            e_var.set(f"{raw[:4]}.{raw[4:6]}.{raw[6:8]}")
                    entry.bind("<KeyRelease>", _auto_date)

                # ---------- 自动格式化：比例 中文冒号 → 英文冒号 ----------
                if fname == "scale":
                    def _auto_scale(evt, e_var=var):
                        val = e_var.get()
                        if "：" in val:
                            e_var.set(val.replace("：", ":"))
                    entry.bind("<KeyRelease>", _auto_scale)

                entries[fname] = (ftype, var)

        # 按钮
        btn_row = tk.Frame(dlg, bg=COLORS["card"])
        btn_row.pack(pady=(36, 28))

        def on_ok():
            nonlocal result
            result = {}
            for fname, cn_name, ftype, required, *_ in fields:
                ftype_key, var = entries[fname]
                val = var.get().strip()
                if val in ("YYYY.MM.DD",):
                    val = ""

                if required and not val:
                    messagebox.showwarning("提示", f"【{cn_name}】为必填项", parent=dlg)
                    return
                if not val:
                    result[fname] = None
                    continue

                if ftype == "decimal":
                    try:
                        result[fname] = float(val)
                    except (ValueError, Exception):
                        messagebox.showwarning("提示", f"【{cn_name}】请输入有效数字", parent=dlg)
                        return
                elif ftype == "date":
                    # 兜底：粘贴的8位数字也自动格式化
                    if len(val) == 8 and val.isdigit():
                        val = f"{val[:4]}.{val[4:6]}.{val[6:8]}"
                        var.set(val)
                    try:
                        parsed = datetime.strptime(val, "%Y.%m.%d").date()
                        if parsed.year < 1900 or parsed.year > 2100:
                            messagebox.showwarning("提示", f"【{cn_name}】请检查日期填写是否正确", parent=dlg)
                            return
                        result[fname] = parsed
                    except ValueError:
                        messagebox.showwarning("提示", f"【{cn_name}】请检查日期填写是否正确", parent=dlg)
                        return
                else:
                    # 比例字段：兜底中文冒号转英文冒号
                    if fname == "scale":
                        val = val.replace("：", ":")
                    result[fname] = val

            dlg.destroy()

        def on_cancel():
            nonlocal result
            result = None
            dlg.destroy()

        RoundedButton(btn_row, text="取消", command=on_cancel,
                      bg=COLORS["separator"], fg=COLORS["text"],
                      hover_bg="#D2D2D7", width=100, height=36, radius=7).pack(side="left", padx=8)
        RoundedButton(btn_row, text="确定", command=on_ok,
                      bg=COLORS["accent"], hover_bg=COLORS["accent_hover"],
                      width=100, height=36, radius=7).pack(side="left", padx=8)

        dlg.bind("<Escape>", lambda e: on_cancel())
        dlg.bind("<Return>", lambda e: on_ok())
        self.wait_window(dlg)
        return result

    # ─── 增删改 ────────────────────────────────
    def _add_record(self):
        table = self._get_table()
        cn = SCHEMA[table]["cn"]
        data = self._form_dialog(f"新增{cn}", table)
        if data is None:
            return

        fields = [k for k, v in data.items() if v is not None]
        if not fields:
            return
        placeholders = ", ".join(f"`{f}`" for f in fields)
        vals = [data[f] for f in fields]

        cur = self._conn.cursor()
        cur.execute(f"INSERT INTO `{table}` ({placeholders}) VALUES ({', '.join(['?']*len(vals))})", vals)
        self._conn.commit()
        new_id = cur.lastrowid
        cur.close()

        self._refresh_table()

        # 新增后提示是否添加图片
        if messagebox.askyesno("添加图片", "是否为这条记录添加图片？"):
            # 选中新插入的行
            for item in self.tree.get_children():
                tags = self.tree.item(item, "tags")
                if tags and str(tags[0]) == str(new_id):
                    self.tree.selection_set(item)
                    self.tree.see(item)
                    break
            self._change_image()

    def _get_selected_id(self):
        sel = self.tree.selection()
        if not sel:
            return None
        tags = self.tree.item(sel[0], "tags")
        return int(tags[0]) if tags else None

    def _edit_record(self):
        table = self._get_table()
        cn = SCHEMA[table]["cn"]
        rid = self._get_selected_id()
        if rid is None:
            return

        cur = self._conn.cursor()
        cur.execute(f"SELECT * FROM `{table}` WHERE id = ?", (rid,))
        row = cur.fetchone()
        cur.close()

        if not row:
            messagebox.showerror("错误", "记录不存在")
            return

        data = self._form_dialog(f"编辑{cn}", table, initial=dict(row))
        if data is None:
            return

        # 空编辑直接返回
        if not data:
            return

        sets = ", ".join(f"`{f}` = ?" for f in data)
        vals = list(data.values()) + [rid]

        cur = self._conn.cursor()
        cur.execute(f"UPDATE `{table}` SET {sets} WHERE id = ?", vals)
        self._conn.commit()
        cur.close()

        self._refresh_table()

    def _delete_record(self):
        table = self._get_table()
        cn = SCHEMA[table]["cn"]
        rid = self._get_selected_id()
        if rid is None:
            return

        # 删除确认弹窗
        confirm = tk.Toplevel(self)
        confirm.title("")
        confirm.geometry("340x160")
        confirm.resizable(False, False)
        confirm.transient(self)
        confirm.grab_set()
        confirm.configure(bg=COLORS["bg"])

        # 先隐藏窗口，计算好位置后再显示，避免左上角闪烁
        confirm.withdraw()
        confirm.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 340) // 2
        y = self.winfo_y() + (self.winfo_height() - 160) // 2
        confirm.geometry(f"+{max(x,0)}+{max(y,0)}")
        confirm.deiconify()

        tk.Label(confirm, text="确认删除", font=(FONT, 15, "bold"),
                 bg=COLORS["bg"], fg=COLORS["text"]).pack(pady=(20, 4))
        tk.Label(confirm, text=f"确定删除这条{cn}记录？此操作不可撤销。",
                 font=(FONT, 11), bg=COLORS["bg"],
                 fg=COLORS["text_secondary"]).pack(pady=(0, 16))

        btn_frame = tk.Frame(confirm, bg=COLORS["bg"])
        btn_frame.pack()

        deleted = {"ok": False}

        def do_delete():
            deleted["ok"] = True
            confirm.destroy()

        RoundedButton(btn_frame, text="取消", command=confirm.destroy,
                      bg=COLORS["separator"], fg=COLORS["text"],
                      hover_bg="#D2D2D7", width=100, height=34, radius=7).pack(side="left", padx=8)
        RoundedButton(btn_frame, text="删除", command=do_delete,
                      bg=COLORS["danger"], hover_bg=COLORS["danger_hover"],
                      width=100, height=34, radius=7).pack(side="left", padx=8)

        confirm.bind("<Escape>", lambda e: confirm.destroy())
        self.wait_window(confirm)

        if not deleted["ok"]:
            return

        # 删除记录并清理关联的图片文件
        cur = self._conn.cursor()
        cur.execute(f"SELECT image_path FROM `{table}` WHERE id = ?", (rid,))
        row = cur.fetchone()
        old_path = row["image_path"] if row else None
        cur.execute(f"DELETE FROM `{table}` WHERE id = ?", (rid,))
        self._conn.commit()
        cur.close()

        # 删除物理图片文件
        if old_path and os.path.isfile(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

        self._show_preview(None)
        self._refresh_table()

    # ─── 导出 Excel ──────────────────────────
    def _export_xlsx(self):
        """导出全部分类的数据到 Excel，每个分类一个 Sheet"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side
        except ImportError:
            messagebox.showerror("缺少依赖",
                                 "导出功能需要 openpyxl 库\n\n请运行:\n  pip install openpyxl",
                                 parent=self)
            return

        # 生成默认文件名
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"Storage_Backup_{ts}.xlsx"

        save_path = filedialog.asksaveasfilename(
            title="导出全部数据",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            parent=self,
        )
        if not save_path:
            return

        # 样式（共用）
        header_font = XlFont(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        cell_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D2D2D7"),
            right=Side(style="thin", color="D2D2D7"),
            top=Side(style="thin", color="D2D2D7"),
            bottom=Side(style="thin", color="D2D2D7"),
        )
        even_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")

        cur = self._conn.cursor()

        wb = Workbook()
        first_sheet = True
        total_records = 0
        sheet_summary = []

        for table in TABLE_ORDER:
            schema = SCHEMA[table]
            cn = schema["cn"]

            # 查询数据
            cur.execute(f"SELECT * FROM `{table}` ORDER BY `id`")
            rows = cur.fetchall()

            # 构建表头（不含 image_path，加入 created_at）
            display_fields = []
            for f in schema["fields"]:
                if f[0] in ("id", "image_path"):
                    continue
                display_fields.append(f)
            display_fields.append(("created_at", "创建时间", "date", False))

            # 创建 Sheet（第一个用 active，后续新建）
            if first_sheet:
                ws = wb.active
                ws.title = cn[:31]
                first_sheet = False
            else:
                ws = wb.create_sheet(title=cn[:31])

            # 写表头
            for col_idx, (_, cn_name, *_) in enumerate(display_fields, 1):
                cell = ws.cell(row=1, column=col_idx, value=cn_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # 写数据
            for row_idx, row in enumerate(rows, 2):
                for col_idx, (fname, _, ftype, *_) in enumerate(display_fields, 1):
                    val = row[fname]
                    if val is None:
                        cell_val = ""
                    elif isinstance(val, (int, float)):
                        cell_val = float(val)
                    elif hasattr(val, "strftime"):
                        cell_val = val
                    else:
                        cell_val = str(val)

                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                    cell.alignment = cell_align
                    cell.border = thin_border
                    if ftype == "date" and isinstance(cell_val, date):
                        cell.number_format = "YYYY.MM.DD"
                    if row_idx % 2 == 0:
                        cell.fill = even_fill

            # 自适应列宽
            for col_idx, (_, cn_name, *_) in enumerate(display_fields, 1):
                max_len = len(cn_name)
                for row_idx in range(2, len(rows) + 2):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val:
                        max_len = max(max_len, len(str(cell_val)))
                ws.column_dimensions[
                    ws.cell(row=1, column=col_idx).column_letter
                ].width = min(max_len + 6, 30)

            # 冻结首行
            ws.freeze_panes = "A2"

            total_records += len(rows)
            if rows:
                sheet_summary.append(f"{cn}: {len(rows)} 条")

        cur.close()

        if total_records == 0:
            messagebox.showinfo("提示", "所有分类都没有数据可导出", parent=self)
            return

        # 保存
        try:
            wb.save(save_path)
        except Exception as e:
            messagebox.showerror("导出失败", f"保存文件时出错:\n{e}", parent=self)
            return

        summary = "\n".join(sheet_summary)
        messagebox.showinfo("导出成功",
                            f"共 {total_records} 条记录\n\n{summary}\n\n{save_path}",
                            parent=self)


if __name__ == "__main__":
    app = App()
    app.mainloop()
