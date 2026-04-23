# -*- coding: utf-8 -*-
"""业务服务：图片管理与 Excel 导出"""

import os
import shutil
import tempfile
import zipfile
from datetime import date

from constants import PICTURES_DIR, unpack_field


class ImageService:
    def __init__(self, pictures_dir, image_folders):
        self._pictures_dir = pictures_dir
        self._image_folders = image_folders

    def ensure_picture_dirs(self):
        for folders in self._image_folders.values():
            for folder_name in folders.values():
                os.makedirs(os.path.join(self._pictures_dir, folder_name), exist_ok=True)

    def get_image_folder(self, table, category=None):
        folders = self._image_folders.get(table, {})
        if category and category in folders:
            folder_name = folders[category]
        elif "_default" in folders:
            folder_name = folders["_default"]
        else:
            folder_name = list(folders.values())[0] if folders else "misc pic"
            os.makedirs(os.path.join(self._pictures_dir, folder_name), exist_ok=True)
        return os.path.join(self._pictures_dir, folder_name)

    @staticmethod
    def make_image_filename(brand, model):
        brand = (brand or "").strip()
        model = (model or "").strip()
        parts = [part for part in (brand, model) if part]
        return "_".join(parts) if parts else "untitled"

    def replace_image(self, table, row, src_path):
        from constants import resolve_image_path

        category = row.get("category") or None
        brand = row.get("brand") or row.get("real_brand") or row.get("model_brand") or ""
        model = row.get("model") or row.get("real_model") or ""

        folder = self.get_image_folder(table, category)
        ext = os.path.splitext(src_path)[1].lower() or ".jpg"
        filename = self.make_image_filename(brand, model) + ext
        dst_path = os.path.join(folder, filename)
        # Resolve old path from DB (may be relative or legacy absolute)
        old_path = resolve_image_path(row.get("image_path"))

        if os.path.abspath(dst_path) != os.path.abspath(src_path):
            base, file_ext = os.path.splitext(filename)
            counter = 1
            while os.path.exists(dst_path) and os.path.abspath(dst_path) != os.path.abspath(old_path or ""):
                filename = f"{base}_{counter}{file_ext}"
                dst_path = os.path.join(folder, filename)
                counter += 1
            shutil.copy2(src_path, dst_path)

        self.remove_image_file(old_path, keep_path=dst_path)
        # Store relative path (relative to PICTURES_DIR) for portability
        return os.path.relpath(dst_path, PICTURES_DIR)

    @staticmethod
    def remove_image_file(image_path, keep_path=None):
        if not image_path or not os.path.isfile(image_path):
            return
        if keep_path and os.path.abspath(image_path) == os.path.abspath(keep_path):
            return
        try:
            os.remove(image_path)
        except OSError:
            pass


def export_all_to_xlsx(save_path, table_order, schema_map, fetch_rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side
    except ImportError:
        return {"error": "missing_openpyxl", "total_records": 0, "sheet_summary": [], "save_path": save_path}

    header_font = XlFont(name="PingFang SC", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = XlFont(name="PingFang SC", size=11)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D2D2D7"),
        right=Side(style="thin", color="D2D2D7"),
        top=Side(style="thin", color="D2D2D7"),
        bottom=Side(style="thin", color="D2D2D7"),
    )

    wb = Workbook()
    first_sheet = True
    total_records = 0
    sheet_summary = []

    for table in table_order:
        schema = schema_map[table]
        cn = schema["cn"]
        rows = fetch_rows(table)

        display_fields = [unpack_field(f) for f in schema["fields"] if unpack_field(f)["name"] not in ("id", "image_path")]
        display_fields.append({"name": "created_at", "cn": "创建时间", "type": "date"})

        if first_sheet:
            ws = wb.active
            ws.title = cn[:31]
            first_sheet = False
        else:
            ws = wb.create_sheet(title=cn[:31])

        for col_idx, f in enumerate(display_fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=f["cn"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            for col_idx, f in enumerate(display_fields, 1):
                fname = f["name"]
                ftype = f["type"]
                val = row[fname]
                if val is None:
                    cell_val = ""
                elif isinstance(val, (int, float)):
                    cell_val = val
                elif hasattr(val, "strftime"):
                    cell_val = val
                else:
                    cell_val = str(val)

                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border
                if ftype == "date" and isinstance(cell_val, date):
                    cell.number_format = "YYYY.MM.DD"

        for col_idx, f in enumerate(display_fields, 1):
            max_len = len(f["cn"])
            for row_idx in range(2, len(rows) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 6, 30)

        ws.freeze_panes = "A2"

        total_records += len(rows)
        if rows:
            sheet_summary.append(f"{cn}: {len(rows)} 条")

    if total_records == 0:
        return {"total_records": 0, "sheet_summary": [], "save_path": save_path}

    wb.save(save_path)
    return {"total_records": total_records, "sheet_summary": sheet_summary, "save_path": save_path}


def _normalize_val(val):
    """Normalize a value for duplicate comparison."""
    if val is None:
        return ""
    s = str(val).strip()
    # Normalize decimal: remove trailing zeros
    try:
        from decimal import Decimal
        d = Decimal(s)
        s = str(d.normalize())
    except Exception:
        pass
    return s.lower()


def import_all_from_xlsx(src_path, table_order, schema_map, insert_fn, list_fn=None):
    """Import records from an xlsx file (same format as export).
    
    Args:
        list_fn: Optional function(table) -> rows to fetch existing records for dedup.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "missing_openpyxl", "total_records": 0}
    from decimal import Decimal

    try:
        wb = load_workbook(src_path, read_only=True, data_only=True)
    except Exception:
        return {"error": "invalid_file", "total_records": 0}

    # Build reverse lookup: sheet cn → table name
    cn_to_table = {}
    for table in table_order:
        cn = schema_map[table]["cn"]
        cn_to_table[cn[:31]] = table
        cn_to_table[cn] = table

    total_records = 0
    skipped = 0
    errors = []

    for ws in wb.worksheets:
        sheet_title = ws.title
        table = cn_to_table.get(sheet_title)
        if not table:
            # Try matching by stripping potential truncation
            for cn, t in cn_to_table.items():
                if sheet_title.startswith(cn[:31]):
                    table = t
                    break
        if not table:
            continue

        schema = schema_map[table]
        all_field_names = {unpack_field(f)["name"] for f in schema["fields"]}
        display_fields = [unpack_field(f) for f in schema["fields"] if unpack_field(f)["name"] not in ("id", "image_path")]
        display_fields.append({"name": "created_at", "cn": "创建时间", "type": "date"})

        # Read header row to map columns
        rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        cn_to_name = {}
        for f in display_fields:
            cn_to_name[f["cn"]] = f["name"]

        col_map = {}  # col_idx → field_name
        for col_idx, cell_val in enumerate(header_row):
            if cell_val and str(cell_val).strip() in cn_to_name:
                col_map[col_idx] = cn_to_name[str(cell_val).strip()]

        if not col_map:
            continue

        # Get field types for conversion
        field_types = {}
        for f in schema["fields"]:
            field = unpack_field(f)
            field_types[field["name"]] = field["type"]

        # Build existing record signatures for dedup
        existing_signatures = set()
        compare_fields = [f for f in display_fields if f["name"] != "created_at"]
        if list_fn:
            try:
                existing_rows = list_fn(table)
                for er in existing_rows:
                    sig_parts = []
                    for f in compare_fields:
                        sig_parts.append(_normalize_val(er[f["name"]]))
                    existing_signatures.add("|".join(sig_parts))
            except Exception:
                pass

        # Read data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            record = {}
            for col_idx, fname in col_map.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None and str(val).strip():
                        # Convert numeric values to Decimal for decimal fields
                        ftype = field_types.get(fname, "text")
                        if ftype == "decimal" and isinstance(val, (int, float)):
                            try:
                                val = Decimal(str(val))
                            except Exception:
                                pass
                        record[fname] = val
                    else:
                        record[fname] = None

            # Only insert if there's meaningful data
            if any(v is not None for k, v in record.items() if k != "created_at"):
                # Check for duplicate
                sig_parts = []
                for f in compare_fields:
                    sig_parts.append(_normalize_val(record.get(f["name"])))
                sig = "|".join(sig_parts)
                if sig in existing_signatures:
                    skipped += 1
                    continue
                existing_signatures.add(sig)

                try:
                    # Filter out fields not in schema
                    filtered = {k: v for k, v in record.items() if k in all_field_names}
                    insert_fn(table, filtered)
                    total_records += 1
                except Exception as e:
                    errors.append(f"{sheet_title} 第{row_idx}行: {e}")

    wb.close()
    return {"total_records": total_records, "errors": errors, "skipped": skipped}


def export_to_zip(zip_path, table_order, schema_map, fetch_rows, pictures_dir=PICTURES_DIR):
    """Export all data and images to a ZIP file.
    
    ZIP structure:
    - data.xlsx (all records)
    - images/{table}/{filename} (image files)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side
    except ImportError:
        return {"error": "missing_openpyxl", "total_records": 0}

    header_font = XlFont(name="PingFang SC", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = XlFont(name="PingFang SC", size=11)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D2D2D7"),
        right=Side(style="thin", color="D2D2D7"),
        top=Side(style="thin", color="D2D2D7"),
        bottom=Side(style="thin", color="D2D2D7"),
    )

    total_records = 0
    image_count = 0

    with tempfile.TemporaryDirectory() as tmp_dir:
        xlsx_path = os.path.join(tmp_dir, "data.xlsx")
        images_tmp_dir = os.path.join(tmp_dir, "images")
        os.makedirs(images_tmp_dir, exist_ok=True)

        wb = Workbook()
        first_sheet = True

        for table in table_order:
            schema = schema_map[table]
            cn = schema["cn"]
            rows = fetch_rows(table)

            display_fields = [unpack_field(f) for f in schema["fields"] if unpack_field(f)["name"] not in ("id", "image_path")]
            display_fields.append({"name": "created_at", "cn": "创建时间", "type": "date"})

            if first_sheet:
                ws = wb.active
                ws.title = cn[:31]
                first_sheet = False
            else:
                ws = wb.create_sheet(title=cn[:31])

            for col_idx, f in enumerate(display_fields, 1):
                cell = ws.cell(row=1, column=col_idx, value=f["cn"])
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            for row_idx, row in enumerate(rows, 2):
                for col_idx, f in enumerate(display_fields, 1):
                    fname = f["name"]
                    ftype = f["type"]
                    val = row[fname]
                    if val is None:
                        cell_val = ""
                    elif isinstance(val, (int, float)):
                        cell_val = val
                    elif hasattr(val, "strftime"):
                        cell_val = val
                    else:
                        cell_val = str(val)

                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                    if ftype == "date" and isinstance(cell_val, date):
                        cell.number_format = "YYYY.MM.DD"

            for col_idx, f in enumerate(display_fields, 1):
                max_len = len(f["cn"])
                for row_idx in range(2, len(rows) + 2):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val:
                        max_len = max(max_len, len(str(cell_val)))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 6, 30)

            ws.freeze_panes = "A2"
            total_records += len(rows)

            # Copy images for this table
            table_img_dir = os.path.join(images_tmp_dir, table)
            os.makedirs(table_img_dir, exist_ok=True)
            for row in rows:
                img_path = row.get("image_path")
                if img_path:
                    # Resolve relative path
                    if not os.path.isabs(img_path):
                        full_path = os.path.join(pictures_dir, img_path)
                    else:
                        full_path = img_path
                    if os.path.isfile(full_path):
                        dst_name = os.path.basename(full_path)
                        dst_path = os.path.join(table_img_dir, dst_name)
                        # Handle duplicate filenames
                        base, ext = os.path.splitext(dst_name)
                        counter = 1
                        while os.path.exists(dst_path):
                            dst_name = f"{base}_{counter}{ext}"
                            dst_path = os.path.join(table_img_dir, dst_name)
                            counter += 1
                        shutil.copy2(full_path, dst_path)
                        image_count += 1

        wb.save(xlsx_path)
        wb.close()

        # Create ZIP file
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(xlsx_path, "data.xlsx")
            for root, dirs, files in os.walk(images_tmp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, tmp_dir)
                    zf.write(file_path, arcname)

    return {"total_records": total_records, "image_count": image_count}


def import_from_zip(zip_path, table_order, schema_map, insert_fn, list_fn=None, pictures_dir=PICTURES_DIR, update_fn=None):
    """Import data and images from a ZIP file.
    
    ZIP structure expected:
    - data.xlsx
    - images/{table}/{filename}
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "missing_openpyxl", "total_records": 0}
    from decimal import Decimal
    from constants import resolve_image_path

    if not zipfile.is_zipfile(zip_path):
        return {"error": "invalid_file", "total_records": 0}

    total_records = 0
    skipped = 0
    errors = []
    image_count = 0

    with tempfile.TemporaryDirectory() as tmp_dir:
        # Extract ZIP
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(tmp_dir)

        xlsx_path = os.path.join(tmp_dir, "data.xlsx")
        images_tmp_dir = os.path.join(tmp_dir, "images")

        if not os.path.isfile(xlsx_path):
            return {"error": "invalid_file", "total_records": 0, "message": "ZIP 中缺少 data.xlsx"}

        # Build reverse lookup: sheet cn → table name
        cn_to_table = {}
        for table in table_order:
            cn = schema_map[table]["cn"]
            cn_to_table[cn[:31]] = table
            cn_to_table[cn] = table

        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception:
            return {"error": "invalid_file", "total_records": 0}

        for ws in wb.worksheets:
            sheet_title = ws.title
            table = cn_to_table.get(sheet_title)
            if not table:
                for cn, t in cn_to_table.items():
                    if sheet_title.startswith(cn[:31]):
                        table = t
                        break
            if not table:
                continue

            schema = schema_map[table]
            all_field_names = {unpack_field(f)["name"] for f in schema["fields"]}
            display_fields = [unpack_field(f) for f in schema["fields"] if unpack_field(f)["name"] not in ("id", "image_path")]
            display_fields.append({"name": "created_at", "cn": "创建时间", "type": "date"})

            rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue

            cn_to_name = {}
            for f in display_fields:
                cn_to_name[f["cn"]] = f["name"]

            col_map = {}
            for col_idx, cell_val in enumerate(header_row):
                if cell_val and str(cell_val).strip() in cn_to_name:
                    col_map[col_idx] = cn_to_name[str(cell_val).strip()]

            if not col_map:
                continue

            field_types = {}
            for f in schema["fields"]:
                field = unpack_field(f)
                field_types[field["name"]] = field["type"]

            existing_signatures = set()
            compare_fields = [f for f in display_fields if f["name"] != "created_at"]
            if list_fn:
                try:
                    existing_rows = list_fn(table)
                    for er in existing_rows:
                        sig_parts = []
                        for f in compare_fields:
                            sig_parts.append(_normalize_val(er[f["name"]]))
                        existing_signatures.add("|".join(sig_parts))
                except Exception:
                    pass

            # Table image folder in Pictures
            table_img_folder = os.path.join(pictures_dir, table)
            os.makedirs(table_img_folder, exist_ok=True)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                record = {}
                for col_idx, fname in col_map.items():
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val is not None and str(val).strip():
                            ftype = field_types.get(fname, "text")
                            if ftype == "decimal" and isinstance(val, (int, float)):
                                try:
                                    val = Decimal(str(val))
                                except Exception:
                                    pass
                            record[fname] = val
                        else:
                            record[fname] = None

                if any(v is not None for k, v in record.items() if k != "created_at"):
                    sig_parts = []
                    for f in compare_fields:
                        sig_parts.append(_normalize_val(record.get(f["name"])))
                    sig = "|".join(sig_parts)
                    if sig in existing_signatures:
                        skipped += 1
                        continue
                    existing_signatures.add(sig)

                    try:
                        filtered = {k: v for k, v in record.items() if k in all_field_names}
                        new_id = insert_fn(table, filtered)
                        total_records += 1

                        # Try to restore matching image
                        if update_fn:
                            brand = record.get("brand") or record.get("real_brand") or record.get("model_brand") or ""
                            model = record.get("model") or record.get("real_model") or ""
                            if brand or model:
                                img_folder = os.path.join(images_tmp_dir, table)
                                if os.path.isdir(img_folder):
                                    brand = (brand or "").strip()
                                    model = (model or "").strip()
                                    base_name = "_".join([p for p in (brand, model) if p]) if (brand or model) else "untitled"
                                    for f in os.listdir(img_folder):
                                        if f.startswith(base_name):
                                            src_img = os.path.join(img_folder, f)
                                            dst_img = os.path.join(table_img_folder, f)
                                            if os.path.exists(dst_img):
                                                base, ext = os.path.splitext(f)
                                                counter = 1
                                                while os.path.exists(dst_img):
                                                    dst_img = os.path.join(table_img_folder, f"{base}_{counter}{ext}")
                                                    counter += 1
                                            shutil.copy2(src_img, dst_img)
                                            rel_path = os.path.relpath(dst_img, pictures_dir)
                                            try:
                                                update_fn(table, new_id, {"image_path": rel_path})
                                            except Exception:
                                                pass
                                            image_count += 1
                                            break
                    except Exception as e:
                        errors.append(f"{sheet_title} 第{row_idx}行: {e}")

        wb.close()

    return {"total_records": total_records, "errors": errors, "skipped": skipped, "image_count": image_count}
